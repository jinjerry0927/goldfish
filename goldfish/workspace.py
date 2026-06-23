"""워크스페이스 스캐폴딩 — ``goldfish init`` 이 만드는 분석 작업 폴더.

생성 구조::

    <폴더>/
    ├─ _템플릿/거래내역_템플릿.csv   (openpyxl 있으면 .xlsx 도)
    ├─ 리포트/                       (결과 HTML 저장 위치)
    ├─ 사용법.txt
    ├─ 리포트_만들기.bat             (Windows 더블클릭)
    └─ run.command                   (macOS/Linux)
"""
from __future__ import annotations

from pathlib import Path

#: 폴더 일괄 분석 시 입력으로 취급하지 않는 하위 폴더
SKIP_DIRS = frozenset({"_템플릿", "_엔진", "리포트"})

CSV_TEMPLATE = (
    "체결일,종목명,매매구분,수량,단가,거래금액,종목코드,실현손익\n"
    "2026-01-15,삼성전자,매수,10,72000,720000,005930,0\n"
    "2026-02-03,삼성전자,매도,10,75000,750000,005930,30000\n"
    "2026-02-10,카카오,매수,5,48000,240000,035720,0\n"
)

USAGE = """🐠 GoldFish 분석 워크스페이스 — 사용법
====================================================

■ 한 줄 요약
  거래내역 파일(CSV 또는 엑셀)을 이 폴더에 넣고
  "리포트_만들기.bat" 를 더블클릭하면 끝! (Mac 은 run.command)

■ 폴더 구성
  리포트_만들기.bat   ← 더블클릭하면 분석 실행 (Windows)
  run.command         ← Mac/Linux 용
  _템플릿\\            ← 직접 입력할 빈 양식(CSV/엑셀)
  리포트\\             ← 만들어진 HTML 리포트가 여기 저장됨
  사용법.txt           ← 이 파일

■ 사용 순서
  1) 거래내역을 준비합니다.
     (A) _템플릿 의 양식을 열어 직접 입력하거나
     (B) 증권사(KB M-able 등)에서 받은 거래내역 CSV/엑셀 파일
  2) 그 파일을 이 폴더 안에 넣습니다(여러 개 가능).
  3) "리포트_만들기.bat" 를 더블클릭합니다.
  4) 잠시 후 '리포트' 폴더에 HTML 리포트가 생깁니다.

■ 자동으로 처리되는 것
  · 숫자에 든 쉼표(예: 1,234) 자동 제거
  · 한글 인코딩(cp949/UTF-8) 자동 인식
  · 엑셀(xlsx)도 그대로 인식 (pip install openpyxl 필요)
  · 빈 행 자동 제거, 종목코드 앞자리 0 보존

■ 입력 컬럼
  필수: 체결일, 종목명, 매매구분, 수량, 단가, 거래금액
  선택: 종목코드, 실현손익(매도 시 손익)

🐠 Watch your money swim.
"""

BAT_LAUNCHER = """@echo off
chcp 65001 >nul
title GoldFish 리포트 만들기
cd /d "%~dp0"
python -m goldfish.cli "%~dp0"
echo.
pause
"""

SH_LAUNCHER = """#!/bin/sh
# GoldFish 리포트 만들기 (macOS/Linux)
cd "$(dirname "$0")"
python3 -m goldfish.cli "$(dirname "$0")"
"""


def _write_xlsx_template(path: Path) -> bool:
    """엑셀 양식 생성. openpyxl 이 없으면 ``False`` 반환(건너뜀)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return False

    gold = "F5A623"
    headers = ["체결일", "종목명", "매매구분", "수량", "단가", "거래금액", "종목코드", "실현손익"]
    examples = [
        ["2026-01-15", "삼성전자", "매수", 10, 72000, 720000, "005930", 0],
        ["2026-02-03", "삼성전자", "매도", 10, 75000, 750000, "005930", 30000],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=gold)
        cell.alignment = Alignment(horizontal="center")
    for row in examples:
        ws.append(row)
    rows = 60
    for r in range(2, rows + 2):
        ws.cell(row=r, column=6).value = f'=IF(AND(D{r}<>"",E{r}<>""),D{r}*E{r},"")'
        for c in (1, 7):  # 체결일·종목코드 텍스트
            ws.cell(row=r, column=c).number_format = "@"
    # 예시 거래금액은 즉시 보이도록 고정
    ws["F2"], ws["F3"] = 720000, 750000
    dv = DataValidation(type="list", formula1='"매수,매도"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C2:C{rows + 1}")
    for col, w in {"A": 14, "B": 16, "C": 10, "D": 10, "E": 12, "F": 14, "G": 12, "H": 12}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return True


def init_workspace(directory: str | Path) -> Path:
    """``directory`` 에 분석 워크스페이스를 만들고 경로를 반환한다.

    이미 존재하는 파일은 덮어쓰지 않는다(여러 번 실행해도 안전).
    """
    directory = Path(directory)
    tpl = directory / "_템플릿"
    tpl.mkdir(parents=True, exist_ok=True)
    (directory / "리포트").mkdir(parents=True, exist_ok=True)

    csv_tpl = tpl / "거래내역_템플릿.csv"
    if not csv_tpl.exists():
        csv_tpl.write_text(CSV_TEMPLATE, encoding="utf-8-sig")

    xlsx_tpl = tpl / "거래내역_템플릿.xlsx"
    if not xlsx_tpl.exists():
        _write_xlsx_template(xlsx_tpl)

    usage = directory / "사용법.txt"
    if not usage.exists():
        usage.write_text(USAGE, encoding="utf-8")

    bat = directory / "리포트_만들기.bat"
    if not bat.exists():
        bat.write_text(BAT_LAUNCHER, encoding="utf-8")

    sh = directory / "run.command"
    if not sh.exists():
        sh.write_text(SH_LAUNCHER, encoding="utf-8")
        try:
            sh.chmod(0o755)
        except OSError:
            pass

    return directory
